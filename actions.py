from pathlib import Path
import sys
from typing import Any, Hashable, Literal, Optional, Sequence
import requests 
import pandas as pd
import openpyxl as opyxl
from openpyxl import Workbook
from pprint import pprint
from selectolax.parser import HTMLParser
from utils import *
from loguru import logger
import httpx

logger.remove(1)
logger.add(sys.stderr,colorize=True)

excel_file_path:Path|str = "book_to_scrape_analysis.xlsx"

def _get_current_row_index(workbook:Workbook,current_sheet_name:str,file_path):
    sheet = workbook[current_sheet_name]
    
    # cells = sheet['A']
    # pprint(list(sheet.iter_rows()))
    current_row: int= len(list(sheet.iter_rows())[1:]) # type: ignore
    # print(" la liste au depart a la taille de {}".format(current_row))
    if current_row == 0:
        return current_row + 1
    return current_row + 1




def _set_cols_width_on_larged_cell_value(file_path:Path|str,sheet_name:str):
    wb: Workbook = opyxl.load_workbook(excel_file_path)
    sh = wb[sheet_name]

    for colonne in sh.columns: # type: ignore
        max_length = 0
        colonne = [cell for cell in colonne]
        for cell in colonne:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)  # type: ignore
            except:
                pass
        ajustement = (max_length + 2) * 1.2
        sh.column_dimensions[colonne[0].column_letter].width = ajustement # type: ignore
    wb.save(excel_file_path)
    
def _first_writer(file_path,mode,header,sheet_name):
    df = pd.DataFrame({        
        "Book ids":[],
        "Book Titles":[]})
    with pd.ExcelWriter(file_path,mode=mode,engine='openpyxl') as writer:
        df.to_excel(writer, index=True, index_label="ID",sheet_name=sheet_name,engine='openpyxl',na_rep="NA")


def unfamous_books_to_excel_file(df:pd.DataFrame,sheet_name:str, file_path:Path|str,mode:Literal['w','a']='w'):
    if mode == 'a':
        with pd.ExcelWriter(file_path,mode=mode,if_sheet_exists='overlay',engine='openpyxl') as writer:
            df.to_excel(writer, index=True,sheet_name=sheet_name,engine='openpyxl',na_rep="NA",index_label="ID",)
        _set_cols_width_on_larged_cell_value(file_path=file_path,sheet_name=sheet_name)
    else:
        with pd.ExcelWriter(file_path,mode=mode,engine='openpyxl') as writer:
            df.to_excel(writer, index=True, index_label="ID",sheet_name=sheet_name,engine='openpyxl',na_rep="NA")
        _set_cols_width_on_larged_cell_value(file_path=file_path,sheet_name=sheet_name)





def check_threshold_in_categories(threshold:int =5):
    """Display all Categories in Library which have amount of books lower to the
    `threshold` parameter   

    :param threshold: define dangerous threshold, defaults to 5
    :type nb: int
    :return: display the name of the categories which doesn't have enought of books
    :rtype:None
    """
    categories_dict = build_categories_dictionnary()
    with requests.Session() as session:
        for name,link in categories_dict.items():
            book_nb = get_amount_of_books(link,from_base_url=False,session=session)[0]
            if book_nb < threshold:
                print(f'Category:{name} only contains {book_nb} books')


def get_home_page_unfamous_books(stars_threshold:int = 2):
    categories = build_categories_dictionnary()
    data = get_worst_rated_books(threshold=stars_threshold)
    books_ids = [book_id for book_id in data.keys()]
    books_titles = [book_title for book_title in data.values()]
    df = pd.DataFrame({
        "Books Ids":books_ids,
        "Books Titles":books_titles
    })
    unfamous_books_to_excel_file(df=df,sheet_name="home_page_unfamous_books",file_path=excel_file_path)
    print("Worst rated books has been wrote in `worst_rated_books.xlsx` file")


def _add_to_books_categories(categorie_name:str, loops:int,targeted_list:list[str|None]):
    for loop in range(loops):
        targeted_list.append(categorie_name)
    return targeted_list



def get_library_unfamous_books(stars_threshold:int = 2):
    categories = build_categories_dictionnary()
    # _first_writer(file_path=excel_file_path,mode='w',header=["Book id", "Book Title"],sheet_name="library_unfamous_books")
    data = {}
    books_categories = []
    with requests.Session() as session:
        for idx,link_category in enumerate(categories.values(),start=0) :
            values = get_worst_rated_books(from_url_page=link_category,threshold=stars_threshold,kwargs={'session':session})
            if values != {}:
                data.update(values)
            categorie_name = [key for key in categories.keys()][idx]
            books_categories = _add_to_books_categories(categorie_name=categorie_name,loops=len(values),targeted_list=books_categories)
        headers_names = ["Book id", "Book Title"]
        books_ids = [book_id for book_id in data.keys()]
        books_titles = [book_title for book_title in data.values()]
        df = pd.DataFrame({
            "Books Ids":books_ids,
            "Book Titles":books_titles,
            "Categories" : books_categories
        })
        unfamous_books_to_excel_file(df=df,  sheet_name="library_unfamous_books",file_path=excel_file_path, mode='a')
        
        # print("Worst rated books has been wrote in `library_unfamous_books.xlsx` file")

def get_book_to_scrape_value(url:str=BASE_URL):
    links = []
    titles_prices_stocks:list[dict[str,str] ]= []
    logger.info("Building details page links list")
    for link in get_books_details_page_link(url):
        links.append(link)
    links = [link if 'catalogue' in link else "catalogue/"+link for link in links]
    logger.info("Building dictionnary with all details price of each book this could take a while ...")
    with httpx.Client() as session:
        for link in links:
            link = urljoin(BASE_URL,link)
            logger.info(f"Get book details  at : {link}")
            response = session.get(link)
            response.raise_for_status()
            tree = HTMLParser(response.content)
            titles_prices_stocks.append({
                'title':get_book_title(tree),
                'price': get_book_price(tree),
                "in_stock": get_in_stock(tree)
            })
    # pprint(titles_prices_stocks)
    total_value = 0
    for book in titles_prices_stocks:
        value_in_stock = float(book["in_stock"]) * float(book["price"])
        print(f"title :{book['title']}| in_stock:{book['in_stock']}|price:${book['price']}| Value in stock : ${value_in_stock} ")
        total_value += value_in_stock
    logger.success(f"The total VALUE OF THIS WEB LIBRARY IS :${total_value}")


    


if __name__ == '__main__':
    pass